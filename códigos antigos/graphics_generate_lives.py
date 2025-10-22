import os
import glob
import openpyxl
import matplotlib
import warnings
warnings.filterwarnings('ignore')
import seaborn as sns
sns.set_style('whitegrid')
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import dataframe_image as dfi
from matplotlib.ticker import MaxNLocator
from datetime import datetime,date
from openpyxl import load_workbook


def buscar_lista_artistas():
    # Acesso exports.txt para buscar o nome dos artistas
    lines = ''
    with open('dados_full/exports.txt') as f:
        lines = f.readlines()
    lines = [i.rstrip() for i in lines]
    return lines

def custom_format(num):
    # Converter para int para evitar decimais
    num = int(num)
    # Formatar com separador de milhares usando ponto
    return f"{num:,}".replace(',', '.')


def dec_format(num):
    if isinstance(num, (int, float)):  # Verificar se o elemento é numérico
        return f'{num:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return str(num)  # Converter qualquer outro tipo para string sem formatação



def format_currency(value):
    # Converter para float e formatar como moeda com 2 casas decimais
    # Usar replace para mudar o separador de milhares para ponto e de decimais para vírgula
    return f'${float(value):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def value_with_arrow(value):
    # Remover o caractere '%' e converter para float
    numeric_value = float(value.rstrip('%'))
    
    # Determina a cor e a direção da seta baseada no valor
    if numeric_value < 0:
        return f"{numeric_value:.2f}% ↓", 'red'
    else:
        return f"{numeric_value:.2f}% ↑", 'green'
    

def quadrant_chart(artista):
    file_path = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'
    data = pd.read_excel(file_path, sheet_name=0)
    data2 = pd.read_excel(file_path, sheet_name=1)

    # Aplicar formatos customizados
    data.iloc[:, 7] = pd.to_numeric(data.iloc[:, 7], errors='coerce')  # Converte para numérico, tornando não numéricos em NaN
    data.iloc[7, 7] = custom_format(data.iloc[7, 7])
    data.iloc[5, 7] = format_currency(data.iloc[5, 7])
    data.iloc[22, 7] = format_currency(data.iloc[22, 7])

    valores = [
        data.iloc[1, 7],
        data.iloc[5, 7],
        data.iloc[7, 7],
        data.iloc[22, 7]
    ]

    valores2 = [
        value_with_arrow(data2.iloc[1, 6]),
        value_with_arrow(data2.iloc[5, 6]),
        value_with_arrow(data2.iloc[7, 6]),
        value_with_arrow(data2.iloc[22, 6])
    ]

    fig, ax = plt.subplots()
    ax.set_xlim(0, 2)
    ax.set_ylim(0, 2)
    ax.axhline(1, color='grey')
    ax.axvline(1, color='grey')
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_linewidth(2)
        spine.set_edgecolor('grey')
    ax.set_xticks([])
    ax.set_yticks([])

    positions = [(0.5, 1.5), (1.5, 1.5), (0.5, 0.5), (1.5, 0.5)]
    positions2 = [(0.5, 1.3), (1.5, 1.3), (0.5, 0.3), (1.5, 0.3)]
    labels = ["Views", "RPM (USD)", "Inscritos", "Receita (USD)"]
    label_positions = [(0.5, 1.8), (1.5, 1.8), (0.5, 0.8), (1.5, 0.8)]

    for pos, val in zip(positions, valores):
        ax.text(*pos, val, ha='center', va='center', fontsize=25)
    for pos, (val, color) in zip(positions2, valores2):
        ax.text(*pos, val, ha='center', va='center', fontsize=10, color=color)

    for label_pos, label in zip(label_positions, labels):
        ax.text(*label_pos, label, ha='center', va='center', fontsize=10)

    plt.tight_layout()  # Ajusta o layout para minimizar margens
    fig.savefig(f'dados_full/{artista}/plots/Resumo.png', bbox_inches='tight') 


def analyze_initial_updated(artista):

    caminho_arquivo = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'

    try:
        # Leitura do arquivo Excel
        df = pd.read_excel(caminho_arquivo, sheet_name='Resultado', header=None)
    except FileNotFoundError:
        print(f"Erro: O arquivo '{caminho_arquivo}' não foi encontrado.")
        return
    except Exception as e:
        print(f"Erro ao ler o arquivo: {e}")
        return

    # Processamento dos dados
    meses = pd.to_datetime(df.iloc[0, 1:].dropna(), errors='coerce')
    visualizacoes = df.iloc[2, 1:len(meses)+1].astype(float)
    receita_total = df.iloc[23, 1:len(meses)+1].astype(float)

    df_anomalias = pd.DataFrame({
        'Meses': meses.reset_index(drop=True),
        'Visualizacoes': visualizacoes.reset_index(drop=True),
        'ReceitaTotal': receita_total.reset_index(drop=True)
    })

    df_anomalias['VisualizacoesMil'] = df_anomalias['Visualizacoes'] / 1000
    df_anomalias['AnomaliaPositiva'] = (df_anomalias['ReceitaTotal'].diff() > 0) & (df_anomalias['Visualizacoes'].diff() <= 0)
    df_anomalias['AnomaliaNegativa'] = (df_anomalias['Visualizacoes'].diff() > 0) & (df_anomalias['ReceitaTotal'].diff() <= 0)

    # Criação do gráfico
    # Criação do gráfico
    plt.figure(figsize=(12, 6))
    ax1 = plt.gca()
    ax1.set_xlabel('Mês')
    ax1.set_ylabel('Receita Total ($)', color='tab:red')
    linha_receita, = ax1.plot(df_anomalias['Meses'], df_anomalias['ReceitaTotal'], color='tab:red', marker='o', linestyle='-', label='Receita Total')
    ax1.tick_params(axis='y', labelcolor='tab:red')
    ax1.set_ylim(bottom=0)  # Define o limite inferior do eixo y para receita

    ax2 = ax1.twinx()
    ax2.set_ylabel('Visualizações (mil)', color='tab:blue')
    linha_visualizacao, = ax2.plot(df_anomalias['Meses'], df_anomalias['VisualizacoesMil'], color='tab:blue', marker='o', linestyle='-', label='Visualizações')
    ax2.tick_params(axis='y', labelcolor='tab:blue')
    ax2.set_ylim(bottom=0)  # Define o limite inferior do eixo y para visualizações

    ax1.legend(handles=[linha_receita], loc='upper left')
    ax2.legend(handles=[linha_visualizacao], loc='upper right')

    for _, row in df_anomalias.iterrows():
        if row['AnomaliaPositiva']:
            ax1.scatter(row['Meses'], row['ReceitaTotal'], color='green', s=100, edgecolor='black', zorder=5)
        if row['AnomaliaNegativa']:
            ax2.scatter(row['Meses'], row['VisualizacoesMil'], color='red', s=100, edgecolor='black', zorder=5)

    plt.title('Receita x Visualização')
    plt.tight_layout()

    # Ajuste no caminho para salvar o gráfico diretamente no diretório de trabalho
    grafico_caminho = f'dados_full/{artista}/plots/Análise_Inicial.jpg'
    plt.savefig(grafico_caminho)


def relevance_graph(artista):

    caminho_arquivo = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'
   
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name='Resultado', header=None)
    except Exception as e:
        print(f"Erro ao ler o arquivo: {e}")
        return

    # Preparando os dados
    months = pd.to_datetime(df.iloc[0, 1:].dropna(), errors='coerce').dropna().reset_index(drop=True)
    impressions = df.iloc[1, 1:len(months)+1].astype(float) / 1e6
    new_impressions = df.iloc[11, 1:len(months)+1].astype(float) / 1e6

    # Configuração do gráfico
    fig, ax = plt.subplots(figsize=(12, 6))
    bar_width = 0.35
    index = np.arange(len(months))

    bars1 = ax.bar(index, impressions, bar_width, label='Impressões Totais', color='blue')
    bars2 = ax.bar(index + bar_width, new_impressions, bar_width, label='Impressões Novas', color='orange')

    # Adicionando rótulos de dados
    for bar in bars1 + bars2:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}',
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),  # Deslocamento vertical para posicionar o rótulo acima da barra
                    textcoords="offset points",
                    ha='center', va='bottom')

    ax.set_xlabel('Meses')
    ax.set_ylabel('Impressões (milhões)')
    ax.set_title('Impressões de Conteúdo Novo vs. Total de Impressões')
    ax.set_xticks(index + bar_width / 2)
    ax.set_xticklabels([month.strftime("%Y-%m") for month in months], rotation=45)
    ax.legend(loc=(0,1))

    plt.tight_layout()

    # Salvando o gráfico
    output_path_png = f'dados_full/{artista}/plots/Relevancia.jpg'
    plt.savefig(output_path_png)


def conversion_graph(artist):
    caminho_arquivo = f'exports_tabelas/Monthly Abril 2024 ({artist})_Lives.xlsx'

    try:
        # Leitura do arquivo Excel
        df = pd.read_excel(caminho_arquivo, sheet_name='Resultado')
        meses = pd.to_datetime(df.columns[2:], format='%Y-%m')
        ctr = df.iloc[12, 2:].str.rstrip('%').astype(float) / 100.0  # Convertendo para float
        visualizacoes = df.iloc[1, 2:].astype(float)  # Suposição que visualizações já estão em formato numérico

        # Criação do gráfico
        fig, ax = plt.subplots(figsize=(10, 6))
        index = np.arange(len(ctr))
        bars = ax.bar(index, ctr * 100, color='blue', label='CTR (%) Novo')

        # Adicionando as anotações de porcentagem em cada barra
        for bar in bars:
            height = bar.get_height()
            label = f'{height:.2f}%'
            ax.annotate(label, xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3), textcoords="offset points", ha='center', va='bottom')

        # Definindo os rótulos do eixo x
        ax.set_xticks(index)
        ax.set_xticklabels([mes.strftime("%Y-%m") for mes in meses], rotation=45)
        ax.set_ylabel('CTR (%) Novo')
        ax.set_xlabel('Meses')
        ax.set_title('Taxa de cliques: Conteúdo Novo')

        # Adicionando o segundo eixo Y
        ax2 = ax.twinx()
        ax2.plot(index, visualizacoes, color='red', marker='o', label='Visualizações')
        ax2.set_ylabel('Visualizações (Milhões)')
        ax2.set_ylim(0, max(visualizacoes) * 1.1)  # Configurando o limite inferior para 0 e superior para 10% acima do máximo

        # Legenda
        ax.legend(loc=(0,1))
        ax2.legend(loc=(0.85, 1))

        # Ajustando layout e salvando o gráfico
        plt.tight_layout()
        plt.savefig(f'dados_full/{artist}/plots/Conversao.jpg')

    except Exception as e:
        print(f"Erro ao gerar o gráfico: {e}")
        return None


def converter_tempo_para_minutos(tempo_str):
    minutos, segundos = map(int, tempo_str.split(':'))
    return minutos + segundos / 60.0


def quality_graph(artista):
    
    workbook = openpyxl.load_workbook(f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx')
    sheet = workbook['Resultado']
    
    meses = [cell.value for cell in sheet['C1':'H1'][0]]
    tempo_medio_assistido_novo = [converter_tempo_para_minutos(cell.value) for cell in sheet['C15':'H15'][0]]
    porcentagem_media_assistida_novo = [float(cell.value.strip('%')) for cell in sheet['C16':'H16'][0]]
    tamanho_medio_video_novo = [float(cell.value) for cell in sheet['C20':'H20'][0]]
    
    fig, ax1 = plt.subplots(figsize=(12, 7))
    
    positions = list(range(len(meses)))
    bar_width = 0.2
    
    bars1 = ax1.bar([p - bar_width/2 for p in positions], tempo_medio_assistido_novo, width=bar_width, color='skyblue', label='Tempo Médio Assistido (min)')
    bars2 = ax1.bar([p + bar_width/2 for p in positions], tamanho_medio_video_novo, width=bar_width, color='orange', label='Tamanho Médio do Vídeo (min)')
    
    for bar in bars1 + bars2:
        height = bar.get_height()
        ax1.annotate(f'{height:.1f}',
                     xy=(bar.get_x() + bar.get_width() / 2, height),
                     xytext=(0, 3),
                     textcoords="offset points",
                     ha='center', va='bottom')

    ax1.set_ylabel('Minutos')
    ax1.set_xticks(positions)
    ax1.set_xticklabels(meses)
    ax1.set_ylim(bottom=0)  # Ajustando o limite inferior do eixo y das barras
    
    ax2 = ax1.twinx()
    line, = ax2.plot(meses, porcentagem_media_assistida_novo, color='red', marker='o', linewidth=2, label='Porcentagem Média Assistida (%)')
    ax2.set_ylabel('%', color='red')
    ax2.tick_params(axis='y', colors='red')
    ax2.set_ylim(bottom=0)  # Opcional: Ajustando o limite inferior do eixo y da linha

    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc=(0,1), fontsize = 8)

    fig.suptitle('Qualidade dos Vídeos Novos')
    plt.tight_layout()
    plt.savefig(f'dados_full/{artista}/plots/Qualidade.jpg')



def monetization_graph(artista):
    
    workbook = openpyxl.load_workbook(f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx')
    sheet = workbook["Resultado"]
    
    rows = {"RPM (USD)": 7, "RPM (USD) Novo": 17, "Receita Total": 24}
    
    months = [sheet.cell(row=1, column=col).value for col in range(3, 9)]
    
    data = {metric: [sheet.cell(row=row, column=col).value for col in range(3, 9)] for metric, row in rows.items()}

    for key in data:
        clean_data = []
        for val in data[key]:
            try:
                clean_val = float(str(val).replace("$", "").replace(",", "").replace("%", ""))
            except ValueError:
                clean_val = 0
            clean_data.append(clean_val)
        data[key] = clean_data
    
    fig, ax1 = plt.subplots(figsize=(10, 6))
    index = np.arange(len(months))
    bar_width = 0.35

    rects1 = ax1.bar(index - bar_width/2, data["RPM (USD)"], bar_width, label='RPM (USD)', color='blue')
    rects2 = ax1.bar(index + bar_width/2, data["RPM (USD) Novo"], bar_width, label='RPM (USD) Novo', color='red')

    for rect in rects1 + rects2:
        height = rect.get_height()
        ax1.annotate(f'{height:.2f}',
                     xy=(rect.get_x() + rect.get_width() / 2, height),
                     xytext=(0, 3),  
                     textcoords="offset points",
                     ha='center', va='bottom')

    ax1.set_xlabel('Meses')
    ax1.set_ylabel('RPM (USD) / RPM (USD) Novo')
    ax1.set_title('Efeito do RPM (Novo x Total) na Receita (USD)')
    ax1.set_xticks(index)
    ax1.set_xticklabels(months)
    ax1.set_ylim(bottom=0)
    
    ax2 = ax1.twinx()
    ax2.plot(months, data["Receita Total"], marker='o', color='green', label='Receita Total (USD)')
    ax2.set_ylabel('Receita Total (USD)', color='green')
    ax2.tick_params(axis='y', labelcolor='green')
    ax2.set_ylim(bottom=0)

    ax1.legend(loc=(0,1))
    ax2.legend(loc=(0.8,1))

    ax1.grid(False)

    fig.tight_layout()
    plt.savefig(f'dados_full/{artista}/plots/Monetizacao.jpg')


def growth_graph(artista):
    
    workbook = openpyxl.load_workbook(f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx')
    sheet = workbook['Resultado']
    
    # Coletando dados das células especificadas
    meses = [sheet[f'{col}1'].value for col in 'CDEFGH']
    saldo_inscritos = [sheet[f'{col}9'].value for col in 'CDEFGH']
    inscritos_vod = [sheet[f'{col}59'].value for col in 'CDEFGH']
    inscritos_shorts = [sheet[f'{col}60'].value for col in 'CDEFGH']
    inscritos_lives = [sheet[f'{col}61'].value for col in 'CDEFGH']

    def safe_float_convert(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    # Convertendo valores usando safe_float_convert
    saldo_inscritos = np.array([safe_float_convert(v) for v in saldo_inscritos])
    inscritos_vod = np.array([safe_float_convert(v) for v in inscritos_vod])
    inscritos_shorts = np.array([safe_float_convert(v) for v in inscritos_shorts])
    inscritos_lives = np.array([safe_float_convert(v) for v in inscritos_lives])

    # Configurações do gráfico
    fig, ax1 = plt.subplots(figsize=(10, 6))
    bar_width = 0.2
    index = np.arange(len(meses))

    bars1 = ax1.bar(index - bar_width, inscritos_vod / 1000, bar_width, label='VODs', color='#000080')
    bars2 = ax1.bar(index, inscritos_shorts / 1000, bar_width, label='Shorts', color='#1E90FF')
    bars3 = ax1.bar(index + bar_width, inscritos_lives / 1000, bar_width, label='Lives', color='#87CEEB')

    ax1.set_xlabel('Meses')
    ax1.set_ylabel('Número de novos inscritos (Milhares)', color='blue')
    ax1.set_xticks(index)
    ax1.set_xticklabels(meses)
    ax1.tick_params(axis='y', labelcolor='blue')

    # Função para adicionar rótulos nos bares
    def add_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax1.annotate(f'{height:.2f}k',
                         xy=(bar.get_x() + bar.get_width() / 2, height),
                         xytext=(0, 3),  # 3 pontos verticais de deslocamento
                         textcoords="offset points",
                         ha='center', va='bottom', fontsize=8)

    add_labels(bars1)
    add_labels(bars2)
    add_labels(bars3)

    ax2 = ax1.twinx()
    ax2.plot(index, saldo_inscritos / 1000, 'r-o', label='Saldo de Inscritos (Total)')
    ax2.set_ylabel('Número total de inscritos no canal (Milhares)', color='red')
    ax2.tick_params(axis='y', labelcolor='red')
    ax2.set_ylim(0, max(saldo_inscritos) / 500)

    ax1.legend(loc='upper left', fontsize=8, bbox_to_anchor=(0, 1.12))
    ax2.legend(loc='upper right', fontsize=8, bbox_to_anchor=(1, 1.06))
    plt.title('Crescimento: VODs x Shorts x Lives')
    plt.tight_layout()
    plt.savefig(f'dados_full/{artista}/plots/Crescimento.jpg')


def traffic_source_chart(artista):
    
    file_path = (f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx')
    df = pd.read_excel(file_path.format(artista=artista), sheet_name=0, skiprows=35, nrows=9)
    df.set_index('Origem de Tráfego', inplace=True)
    df.drop('Média', axis=1, inplace=True)
    df = df.T 
    df

    # Criando o gráfico
    ax = df.plot(kind='bar', stacked=True, figsize=(12, 7), color=plt.cm.get_cmap('tab20').colors)
    ax.set_title('Principais Origens do Tráfego')
    ax.set_ylabel('Porcentagem (%)')
    ax.set_xlabel('Meses')
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))
    ax.legend(title='Métricas', bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.savefig(f'dados_full/{artista}/plots/Origem_do_trafego.jpg')


def revenue_graph(artista):
    
    workbook = openpyxl.load_workbook(f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx', data_only=True)
    sheet = workbook['Resultado']

    metricas = ['Receita Total', 'Receita % Novo', 'Receita % Velho']
    meses_corretos = [sheet[f'{col}1'].value for col in ['C', 'D', 'E', 'F', 'G', 'H']]

    dados_metricas = {}
    for metrica in metricas:
        linha_metrica = [i+1 for i, valor in enumerate(sheet['A']) if valor.value == metrica][0]
        dados_metricas[metrica] = [sheet[f'{col}{linha_metrica}'].value for col in ['C', 'D', 'E', 'F', 'G', 'H']]

    dados_receita_total_float = [float(valor.replace(',', '').replace('$', '')) for valor in dados_metricas['Receita Total']]
    dados_receita_percentual_novo_float = [float(valor.strip('%')) for valor in dados_metricas['Receita % Novo']]
    dados_receita_percentual_velho_float = [float(valor.strip('%')) for valor in dados_metricas['Receita % Velho']]

    fig, ax1 = plt.subplots(figsize=(14, 8))
    index = np.arange(len(meses_corretos))
    bar_width = 0.2

    cmap = matplotlib.colormaps['tab20']

    bars1 = ax1.bar(index - bar_width/2, dados_receita_percentual_novo_float, bar_width, color=cmap.colors[0], label='% Novo')
    bars2 = ax1.bar(index + bar_width/2, dados_receita_percentual_velho_float, bar_width, color=cmap.colors[1], label='% Velho')

    # Adicionando rótulos de dados nas colunas
    ax1.bar_label(bars1, label_type='center', fmt='%d')
    ax1.bar_label(bars2, label_type='center', fmt='%d')

    ax1.set_xlabel('Mês')
    ax1.set_xticks(index)
    ax1.set_xticklabels(meses_corretos, rotation=45)
    ax1.set_ylabel('% de Receita')
    ax1.set_ylim(0, 100)

    ax2 = ax1.twinx()
    line, = ax2.plot(meses_corretos, dados_receita_total_float, 'r-o', label='Receita Total (USD)')
    
    ax2.set_ylabel('USD')
    ax2.tick_params(axis='y', colors='r')

    # Corrigindo a inclusão da legenda
    handles = [bars1, bars2, line]
    labels = ['% Novo', '% Velho', 'Receita Total (USD)']
    ax1.legend(handles, labels, loc=(0,1), bbox_to_anchor=(0, 1))

    plt.title('Análise da Receita: Novo vs Velho')
    plt.tight_layout()
    plt.savefig(f'dados_full/{artista}/plots/Receita.jpg')


def custom_format(num):
    # Converter para int para evitar decimais
    num = int(num)
    # Formatar com separador de milhares usando ponto
    return f"{num:,}".replace(',', '.')

def format_currency(value):
    # Converter para float e formatar como moeda com 2 casas decimais
    # Usar replace para mudar o separador de milhares para ponto e de decimais para vírgula
    return f'${float(value):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

def value_with_arrow(value):
    # Remover o caractere '%' e converter para float
    numeric_value = float(value.rstrip('%'))
    
    # Determina a cor e a direção da seta baseada no valor
    if numeric_value < 0:
        return f"{numeric_value:.2f}% ↓", 'red'
    else:
        return f"{numeric_value:.2f}% ↑", 'green'


def community_table(artista):
    file_path = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'
    df = pd.read_excel(file_path.format(artista=artista), sheet_name=0, skiprows=45, nrows=5)
    df.drop('Média', axis=1, inplace=True)
    df.set_index('Comunidade', inplace=True)
    df = df.T
    # Convertendo as colunas para formato numérico, onde aplicável
    for col in df.columns[0:]:  # Assumindo que a primeira coluna é de datas
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df.reset_index(inplace=True)
    df['index'] = pd.to_datetime(df['index']).dt.strftime('%b')
    df.rename(columns={'index':'Comunidade'}, inplace=True)
    df.set_index('Comunidade', inplace=True)
    df = df.T
    df.insert(0,'Média', [0,0,0,0,0])

    # Calculando mudanças percentuais para colorir as células
    df_pct = df.T.pct_change().iloc[1:].T
    df_pct.fillna(0, inplace=True)  # Tratar NaNs resultantes de pct_change()

    X = np.array([np.array(['#5FAD56' if j >= 0 else '#EF3E36' for j in i]) for i in df_pct.values])

    # Removendo a primeira coluna de datas após o uso
    df.drop(df.columns[0], axis=1, inplace=True)

    # Preparando o texto das células
    X_text = np.array(df)
    X_text = X_text.tolist()
    
    for i in range(len(X_text)):
        X_text[i] = [f'{int(x):,}'.replace(',', '.') if isinstance(x, (int, float)) else x for x in X_text[i]]
        if i == len(X_text) - 1:
            X_text[i] = [f"{x} %" for x in X_text[i]]
    
    # Criação da tabela visual no Matplotlib
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.set_title('Métricas na Comunidade', fontsize=30, y=1.12)
    ax.axis('off')
    tabela = ax.table(cellText=X_text, colLabels=df.columns, loc='center', cellColours=X, rowLabels=df.index)
    tabela.scale(1, 3)
    tabela.set_fontsize(20)
    fig.savefig(f'dados_full/{artista}/plots/comunidade.jpg', transparent=True, bbox_inches='tight')


def publicated_table(artista):

    file_path = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'
    # Carregar linhas 46 a 51 da primeira aba (linhas em índice do Python começam em 0, então usamos 45 a 50)
    df = pd.read_excel(file_path.format(artista=artista), sheet_name=0, skiprows=51, nrows=3)
    df.set_index('Vídeos Publicados', inplace=True)
    df = df.T
    df.drop('Média', inplace=True)
    df.reset_index(inplace=True)
    df['index'] = pd.to_datetime(df['index']).dt.strftime('%b')
    df = df.T
    df.reset_index(inplace=True)
    header = df.iloc[0]
    df = df[1:]
    df.columns = header
    df.set_index('index', inplace=True)
    df.insert(0, 'Média', [0,0,0])

    X = np.array([np.array(['#87CEEB' if j >= 0 else '#87CEEB' for j in i.tolist()]) for i in np.array(df.T.pct_change().iloc[1:]).T])

    # Removendo a primeira coluna de datas após o uso
    df.drop(df.columns[0], axis=1, inplace=True)

    # Processo para coloração baseado nos valores de mudança percentual
    alpha = 0.5
    h = [[j.lstrip('#') for j in i] for i in X]

    def trans(x):
        return tuple([round(int(x[i:i+2], 16) / 255, 3) for i in (0, 2, 4)] + [alpha])

    for i in range(len(h)):
        for j in range(len(h[i])):
            h[i][j] = trans(h[i][j])
    h = np.array([np.array(i) for i in h])

    # Preparando o texto das células
    X = np.array(df)
    X = X.astype(int)
    X = X.tolist()

    # Criação da tabela visual no Matplotlib
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.set_title('Vídeos Publicados', fontsize=30, y=0.9)
    ax.axis('off')
    tabela = ax.table(cellText=X, colLabels=df.columns, loc='center', cellColours=h, rowLabels=df.index)
    tabela.scale(1, 3)
    tabela.set_fontsize(20)
    fig.savefig(f'dados_full/{artista}/plots/Publicados.png', transparent=True, bbox_inches='tight')


def watch_table(artista):

    file_path = f'exports_tabelas/Monthly Abril 2024 ({artista})_Lives.xlsx'
    # Carregar linhas 46 a 51 da primeira aba (linhas em índice do Python começam em 0, então usamos 45 a 50)
    df = pd.read_excel(file_path.format(artista=artista), sheet_name=0, skiprows=55, nrows=1)
    df = df.T
    df.reset_index(inplace=True)
    df = df.T
    df.iloc[0,2:] = pd.to_datetime(df.iloc[0,2:]).dt.strftime('%b')
    header = df.iloc[0]
    df = df[1:]
    df.columns = header
    df.set_index('WatchTime', inplace=True)
    

    X = np.array([np.array(['#87CEEB' if j >= 0 else '#87CEEB' for j in i.tolist()]) for i in np.array(df.T.pct_change().iloc[1:]).T])

    # Removendo a primeira coluna de datas após o uso
    df.drop(df.columns[0], axis=1, inplace=True)

    # Processo para coloração baseado nos valores de mudança percentual
    alpha = 0.5
    h = [[j.lstrip('#') for j in i] for i in X]

    def trans(x):
        return tuple([round(int(x[i:i+2], 16) / 255, 3) for i in (0, 2, 4)] + [alpha])

    for i in range(len(h)):
        for j in range(len(h[i])):
            h[i][j] = trans(h[i][j])
    h = np.array([np.array(i) for i in h])

    # Preparando o texto das células
    X = np.array(df)
    vectorized_format = np.vectorize(dec_format)
    X = vectorized_format(df)
    X = X.tolist()

    # Criação da tabela visual no Matplotlib
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.set_title('WatchTime', fontsize=30, y=0.7)
    ax.axis('off')
    tabela = ax.table(cellText=X, colLabels=df.columns, loc='center', cellColours=h, rowLabels=df.index)
    tabela.scale(1, 3)
    tabela.set_fontsize(20)
    fig.savefig(f'dados_full/{artista}/plots/Watchtime.png', transparent=True, bbox_inches='tight')


def custom_format(value):
    return "{:,}".format(value)  # Formata números com separadores de milhares

def align_right(v):
    return "text-align: right;"

def align_left(v):
    return "text-align: left;"

def x1(v):
        return "max-width: 35ch;"

def top_vids(artista):
    file_path = f'dados_full/{artista}/topVids.csv'
    df = pd.read_csv(file_path)
    df = df.drop(df.index[0])    
    df = df[['Título do vídeo', 'Visualizações']].sort_values(by='Visualizações', ascending=False).head(10)
    df.iloc[:, 1] = df.iloc[:, 1].apply(custom_format)

    # Aplica estilos específicos para cada coluna
    df = df.style.applymap(align_left, subset=['Título do vídeo']) \
                 .applymap(align_right, subset=['Visualizações']) \
                 .applymap_index(x1)
    dfi.export(df, f'dados_full/{artista}/plots/top_vids.jpg', dpi=800, fontsize=5)


def run():
    lista_de_artistas = buscar_lista_artistas()

    for artista in lista_de_artistas:
        print('Gerando: '+ artista)
        quadrant_chart(artista)
        analyze_initial_updated(artista)
        relevance_graph(artista)
        conversion_graph(artista)
        quality_graph(artista)
        monetization_graph(artista)
        growth_graph(artista)
        traffic_source_chart(artista)
        revenue_graph(artista)
        community_table(artista)
        publicated_table(artista)
        watch_table(artista)
        top_vids(artista)

        print('Done')


if __name__ == "__main__":
    run()